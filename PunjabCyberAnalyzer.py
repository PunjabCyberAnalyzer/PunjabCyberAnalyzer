import streamlit as st
import pandas as pd
import re
import io
import os
import csv
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle as PDFTableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import json
import hashlib
import platform
import tempfile

# ====== RUNTIME DEPENDENCY FIX ======
os.system("pip install streamlit pandas openpyxl reportlab python-docx --quiet")

# Define operator IDs, passwords, and license keys
OPERATOR_IDS = [f"PCO{i:03d}" for i in range(1, 51)]  # PCO001 to PCO050
PASSWORDS = [
    "ZAIN###1234", "AHMED###5678", "KHAN###9012", "ALI###3456", "REHMAN###7890",
    "IQBAL###1234", "HASSAN###5678", "FAROOQ###9012", "YOUSUF###3456", "NADEEM###7890",
    "SAJID###1234", "RAZA###5678", "SHAHID###9012", "TARIQ###3456", "WAQAS###7890",
    "ZUBAIR###1234", "ASIM###5678", "BILAL###9012", "DANISH###3456", "EHSAN###7890",
    "FAISAL###1234", "GHAFOOR###5678", "HAMZA###9012", "IMRAN###3456", "JAVED###7890",
    "KHALID###1234", "LIAQAT###5678", "MAJEED###9012", "NADEEM###3456", "OMAR###7890",
    "QASIM###1234", "RAFIQ###5678", "SAIF###9012", "TAHIR###3456", "USMAN###7890",
    "VASEEM###1234", "WAHEED###5678", "XAVIER###9012", "YASIR###3456", "ZAFAR###7890",
    "AAMIR###1234", "BAKHT###5678", "CHAUDH###9012", "DAWOOD###3456", "EJAZ###7890",
    "FAZAL###1234", "GHULAM###5678", "HABIB###9012", "IRFAN###3456", "JAVED###7890"
]
VALID_KEYS = [
    "PCYBR-PK01-9K7M", "PKCYB-X7J2-4L9P", "CYBPR-M9K4-2L7X", "PRCYB-J2L9-7K4X",
    "BYCPR-X4L7-9M2K", "YCPRB-L9M2-4K7X", "CPRBY-7K4X-2L9M", "RBCPY-M2L9-7K4X",
    "BYCPR-X9L7-2M4K", "PRCYB-J4L9-7K2X", "CYBPR-M7K2-4L9X", "YCPRB-L2M9-7K4X",
    "CPRBY-9K4X-2L7M", "RBCPY-M4L7-9K2X", "BYCPR-X2L9-7M4K", "YCPRB-L7M2-4K9X",
    "CPRBY-2K4X-9L7M", "RBCPY-M9L2-7K4X", "BYCPR-X7L4-2M9K", "PRCYB-J9L7-4K2X",
    "CYBPR-M2K9-7L4X", "YCPRB-L4M7-9K2X", "CPRBY-7L2X-9K4M", "RBCPY-M7L9-2K4X",
    "BYCPR-X4L9-7M2K", "YCPRB-L9M4-2K7X", "CPRBY-2L7X-9K4M", "RBCPY-M4L2-7K9X",
    "BYCPR-X9L2-4M7K", "PRCYB-J7L4-9K2X", "CYBPR-M9K7-2L4X", "YCPRB-L2M7-9K4X",
    "CPRBY-4L9X-7K2M", "RBCPY-M7L2-9K4X", "BYCPR-X2L7-4M9K", "YCPRB-L9M2-7K4X",
    "CPRBY-7L4X-2K9M", "RBCPY-M2L7-9K4X", "BYCPR-X4L9-2M7K", "PRCYB-J9L2-7K4X",
    "CYBPR-M7K9-2L4X", "YCPRB-L4M9-7K2X", "CPRBY-9L2X-7K4M", "RBCPY-M7L4-9K2X",
    "BYCPR-X9L7-2M4K", "YCPRB-L2M4-7K9X", "CPRBY-4L7X-9K2M", "RBCPY-M9L7-2K4X",
    "BYCPR-X7L9-4M2K", "PRCYB-J2L4-7K9X"
]

st.set_page_config(layout="wide")
st.markdown("""
<style>
.stApp {
    background-color: #000000 !important;
    background-image:
        linear-gradient(to bottom, #001F00, #000000),
        url('https://upload.wikimedia.org/wikipedia/commons/thumb/3/32/Flag_of_Pakistan.svg/2560px-Flag_of_Pakistan.svg.png') !important;
    background-size: contain;
    background-repeat: no-repeat;
    background-position: center;
    font-family: 'Courier New', monospace !important;
    opacity: 0.95;
}
.top-header {
    background: linear-gradient(90deg, #001F00, #006400, #001F00);
    height: 80px;
    border-bottom: 3px solid #00FF00;
    margin-bottom: 20px;
    box-shadow: 0 0 15px #00FF00;
}
.header-left {
    color: #00FF00 !important;
    position: fixed;
    left: 10px;
    top: 25px;
    font-size: 1.5rem;
    text-shadow: 0 0 5px #00FF00;
    font-weight: bold;
}
.header-right {
    color: #00FF00 !important;
    position: fixed;
    right: 10px;
    top: 25px;
    font-size: 1.5rem;
    text-shadow: 0 0 5px #00FF00;
    font-weight: bold;
}
.name-header-sparkle {
    color: #00FF00 !important;
    text-align: center;
    font-size: 2.8rem;
    text-shadow: 0 0 15px #00FF00, 0 0 25px #00FF00, 0 0 35px #00FF00;
    margin: 1rem 0;
    letter-spacing: 3px;
    font-weight: bolder;
    font-family: 'Courier New', monospace;
    position: relative;
    animation: neonGlow 1.5s ease-in-out infinite alternate;
}
.name-header-sparkle::after {
    content: '✨';
    position: absolute;
    color: #00FF00;
    font-size: 1.5rem;
    top: -10px;
    right: -30px;
    animation: sparkle 2s ease-in-out infinite;
}
.name-header-sparkle small {
    display: block;
    font-size: 1.2rem;
    color: #00FF00;
    text-shadow: 0 0 10px #00FF00;
    letter-spacing: 2px;
    margin-top: 5px;
}
@keyframes neonGlow {
    from { text-shadow: 0 0 10px #00FF00, 0 0 20px #00FF00, 0 0 30px #00FF00; }
    to { text-shadow: 0 0 20px #00FF00, 0 0 30px #00FF00, 0 0 40px #00FF00; }
}
@keyframes sparkle {
    0% { opacity: 0; transform: scale(0.5); }
    50% { opacity: 1; transform: scale(1.2); }
    100% { opacity: 0; transform: scale(0.5); }
}
.whatsapp-container {
    text-align: center;
    margin: 5px auto 10px auto;
    width: 80%;
    border: 1px solid #00FF00;
    padding: 12px;
    background: linear-gradient(135deg, #001F00 0%, #006400 50%, #001F00 100%);
    box-shadow: 0 0 15px #00FF00, inset 0 0 10px #00FF00;
    position: relative;
    overflow: hidden;
}
.whatsapp-link {
    font-family: 'Courier New';
    font-size: 1.1rem;
    color: #00FF00;
    text-shadow: 0 0 5px #00FF00;
    text-decoration: none;
    display: block;
    position: relative;
    z-index: 2;
}
.whatsapp-container::before {
    content: "";
    position: absolute;
    top: -100%;
    left: -100%;
    width: 300%;
    height: 300%;
    background: linear-gradient(
        to right,
        rgba(0, 255, 0, 0) 0%,
        rgba(0, 255, 0, 0.3) 50%,
        rgba(0, 255, 0, 0) 100%
    );
    transform: rotate(45deg);
    animation: tacticalShine 3.5s infinite;
    z-index: 1;
}
.notice-container {
    text-align: center;
    margin: 10px auto;
    width: 80%;
    border: 1px solid #00FF00;
    padding: 8px;
    background: linear-gradient(135deg, #001F00 0%, #006400 50%, #001F00 100%);
    box-shadow: 0 0 10px #00FF00, inset 0 0 5px #00FF00;
    position: relative;
    overflow: hidden;
}
.notice-text {
    font-family: 'Courier New';
    font-size: 0.9rem;
    color: #00FF00;
    text-shadow: 0 0 3px #00FF00;
    font-weight: bold;
}
.notice-container::before {
    content: "";
    position: absolute;
    top: -100%;
    left: -100%;
    width: 300%;
    height: 300%;
    background: linear-gradient(
        to right,
        rgba(0, 255, 0, 0) 0%,
        rgba(0, 255, 0, 0.2) 50%,
        rgba(0, 255, 0, 0) 100%
    );
    transform: rotate(45deg);
    animation: tacticalShine 4s infinite;
    z-index: 0;
}
.email-container {
    text-align: center;
    margin: 10px auto;
    width: 80%;
    border: 1px solid #00FF00;
    padding: 12px;
    background: linear-gradient(135deg, #001F00 0%, #006400 50%, #001F00 100%);
    box-shadow: 0 0 15px #00FF00, inset 0 0 10px #00FF00;
    position: relative;
    overflow: hidden;
}
.email-link {
    font-family: 'Courier New';
    font-size: 1.1rem;
    color: #00FF00;
    text-shadow: 0 0 5px #00FF00;
    text-decoration: none;
    display: block;
    position: relative;
    z-index: 2;
}
.email-container::before {
    content: "";
    position: absolute;
    top: -100%;
    left: -100%;
    width: 300%;
    height: 300%;
    background: linear-gradient(
        to right,
        rgba(0, 255, 0, 0) 0%,
        rgba(0, 255, 0, 0.3) 50%,
        rgba(0, 255, 0, 0) 100%
    );
    transform: rotate(45deg);
    animation: tacticalShine 3.5s infinite;
    z-index: 1;
}
.stButton>button {
    display: block !important;
    margin: 10px auto !important;
    width: 320px !important;
    height: 60px !important;
    border: 3px solid #00FF00 !important;
    background: linear-gradient(135deg, #001F00 0%, #006400 50%, #001F00 100%) !important;
    color: #00FF00 !important;
    font-family: 'Courier New' !important;
    font-size: 1.2rem !important;
    font-weight: bold !important;
    letter-spacing: 2px !important;
    border-radius: 0 !important;
    box-shadow: 0 0 20px #00FF00, inset 0 0 10px #00FF00 !important;
    text-shadow: 0 0 8px #00FF00 !important;
    transition: all 0.3s !important;
    position: relative !important;
    overflow: hidden !important;
    z-index: 1 !important;
}
.stButton>button:hover {
    background: linear-gradient(135deg, #00AA00 0%, #00FF00 50%, #00AA00 100%) !important;
    color: #000 !important;
    box-shadow: 0 0 30px #00FF00, inset 0 0 15px #00FF00 !important;
    transform: scale(1.05) !important;
}
.stButton>button:active {
    transform: scale(0.98) !important;
}
.stButton>button::before {
    content: "";
    position: absolute;
    top: -100%;
    left: -100%;
    width: 300%;
    height: 300%;
    background: linear-gradient(
        to right,
        rgba(0, 255, 0, 0) 0%,
        rgba(0, 255, 0, 0.3) 50%,
        rgba(0, 255, 0, 0) 100%
    );
    transform: rotate(45deg);
    animation: tacticalShine 3.5s infinite;
    z-index: -1;
}
@keyframes tacticalShine {
    0% { left: -100%; top: -100%; }
    100% { left: 100%; top: 100%; }
}
.key-info-container {
    text-align: center;
    margin: 10px auto 20px auto;
    width: 50%;
    border: 1px solid #00FF00;
    padding: 8px;
    background: linear-gradient(135deg, #001F00 0%, #006400 50%, #001F00 100%);
    box-shadow: 0 0 10px #00FF00, inset 0 0 5px #00FF00;
    position: relative;
    overflow: hidden;
}
.key-info-text {
    font-family: 'Courier New';
    font-size: 0.9rem;
    color: #00FF00;
    text-shadow: 0 0 3px #00FF00;
    font-weight: bold;
}
.key-info-container::before {
    content: "";
    position: absolute;
    top: -100%;
    left: -100%;
    width: 300%;
    height: 300%;
    background: linear-gradient(
        to right,
        rgba(0, 255, 0, 0) 0%,
        rgba(0, 255, 0, 0.2) 50%,
        rgba(0, 255, 0, 0) 100%
    );
    transform: rotate(45deg);
    animation: tacticalShine 4s infinite;
    z-index: 0;
}
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="top-header"></div>', unsafe_allow_html=True)
st.markdown('<div class="header-left">PUNJAB POLICE ZINDABAD</div>', unsafe_allow_html=True)
st.markdown('<div class="header-right">PAKISTAN ZINDABAD</div>', unsafe_allow_html=True)
st.markdown('<div class="name-header-sparkle">PUNJAB CYBER ANALYZER<br><small>by Punjab Softwares</small></div>', unsafe_allow_html=True)

def get_device_id():
    """Generate a unique device ID based on system information."""
    try:
        system_info = f"{platform.node()}{platform.system()}{platform.processor()}"
        return hashlib.sha256(system_info.encode()).hexdigest()
    except:
        import uuid
        return str(uuid.uuid4())

def load_key_data():
    """Load key data from JSON file."""
    try:
        if os.path.exists("key_data.json"):
            with open("key_data.json", "r") as f:
                return json.load(f)
        return {}
    except:
        return {}

def save_key_data(key_data):
    """Save key data to JSON file."""
    try:
        with open("key_data.json", "w") as f:
            json.dump(key_data, f, indent=4)
    except:
        pass

@st.cache_data
def load_file(uploaded_file):
    """Load and cache file data with enhanced column detection for Excel and CSV."""
    if uploaded_file is None:
        return None
    try:
        if uploaded_file.name.endswith((".xlsx", ".xls")):
            df = pd.read_excel(uploaded_file, dtype=str, engine='openpyxl')
        else:
            encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'utf-16', 'windows-1252']
            delimiters = [',', ';', '\t']
            uploaded_file.seek(0)
            for encoding in encodings:
                for delimiter in delimiters:
                    try:
                        uploaded_file.seek(0)
                        df = pd.read_csv(
                            uploaded_file,
                            dtype=str,
                            sep=delimiter,
                            encoding=encoding,
                            on_bad_lines='skip',
                            quoting=csv.QUOTE_ALL,
                            escapechar='\\'
                        )
                        if not df.empty:
                            break
                    except Exception:
                        continue
                else:
                    continue
                break
            else:
                st.error("Failed to parse CSV file with available encodings and delimiters. Check file format.")
                return None

        columns = []
        seen = {}
        for col in df.columns:
            col_clean = col.strip()
            if col_clean in seen:
                seen[col_clean] += 1
                columns.append(f"{col_clean}_{seen[col_clean]}")
            else:
                seen[col_clean] = 0
                columns.append(col_clean)
        df.columns = columns

        def is_phone_number(value):
            if pd.isna(value):
                return False
            value = str(value).strip()
            return bool(re.match(r'^(?:03\d{9}|(?:\+|00)923\d{9})$', value))

        def is_imei(value):
            if pd.isna(value):
                return False
            value = str(value).strip()
            return bool(re.match(r'^\d{14,16}$', value))

        def is_call_type(value):
            if pd.isna(value):
                return False
            value = str(value).strip().lower()
            return value in ['incoming', 'outgoing', 'missed', 'sms', 'data', 'voice', 'in', 'out']

        def is_location(value):
            if pd.isna(value):
                return False
            value = str(value).strip()
            return bool(re.match(r'^[A-Za-z0-9\s,.-]+$', value)) and len(value) <= 100 and not value.isdigit()

        def is_mixed_type(column):
            try:
                types = set(df[column].dropna().apply(lambda x: isinstance(x, str) and not is_phone_number(x)))
                return len(types) > 1 or any(not is_phone_number(x) for x in df[column].dropna())
            except:
                return False

        column_map = {
            'A Number': 'A-Party', 'B Number': 'B-Party', 'Phone': 'A-Party', 'Mobile': 'A-Party',
            'Number': 'A-Party', 'Type': 'Call Type', 'Direction': 'Call Type', 'Call_Type': 'Call Type',
            'CallType': 'Call Type', 'Location': 'SiteLocation', 'Site': 'SiteLocation', 'Cell': 'SiteLocation',
            'Cell_Site': 'SiteLocation', 'Area': 'SiteLocation', 'MSISDN': 'A-Party', 'CALL_ORG_NUM': 'B-Party',
            'CELL ID': 'B-Party', 'A PARTY': 'A-Party', 'B PARTY': 'B-Party', 'CALL DETAIL NUMBER': 'B-Party',
            'IMSI': 'IMEI', 'Device_ID': 'IMEI', 'IMEI Number': 'IMEI', 'IMEI_NO': 'IMEI'
        }

        a_party_col = None
        b_party_col = None
        imei_col = None
        call_type_col = None
        location_col = None

        for col in df.columns:
            col_upper = col.strip().upper()
            if col_upper in [k.upper() for k in column_map if column_map[k] == 'A-Party']:
                a_party_col = col
            elif col_upper in [k.upper() for k in column_map if column_map[k] == 'B-Party']:
                b_party_col = col
            elif col_upper in [k.upper() for k in column_map if column_map[k] == 'IMEI']:
                imei_col = col
            elif col_upper in [k.upper() for k in column_map if column_map[k] == 'Call Type']:
                call_type_col = col
            elif col_upper in [k.upper() for k in column_map if column_map[k] == 'SiteLocation']:
                location_col = col
            elif not a_party_col and df[col].head(20).apply(is_phone_number).any():
                a_party_col = col
            elif not b_party_col and is_mixed_type(col):
                b_party_col = col
            elif not imei_col and df[col].head(20).apply(is_imei).any():
                imei_col = col
            elif not call_type_col and df[col].head(20).apply(is_call_type).any():
                call_type_col = col
            elif not location_col and df[col].head(20).apply(is_location).any():
                location_col = col

        if a_party_col and b_party_col:
            same_number_rows = df[df[a_party_col] == df[b_party_col]]
            if not same_number_rows.empty:
                pass

        final_columns = []
        for col in df.columns:
            col_clean = col.strip()
            if col == a_party_col:
                final_columns.append('A-Party')
            elif col == b_party_col:
                final_columns.append('B-Party')
            elif col == imei_col:
                final_columns.append('IMEI')
            elif col == call_type_col:
                final_columns.append('Call Type')
            elif col == location_col:
                final_columns.append('SiteLocation')
            else:
                final_columns.append(column_map.get(col_clean, col_clean))
        df.columns = final_columns

        final_columns = []
        seen_columns = {}
        for col in df.columns:
            if col in seen_columns:
                seen_columns[col] += 1
                final_columns.append(f"{col}_{seen_columns[col]}")
            else:
                seen_columns[col] = 0
                final_columns.append(col)
        df.columns = final_columns

        df = df.reset_index(drop=True)
        return df
    except Exception as e:
        st.error(f"Failed to load file: {e}")
        return None

def create_excel(df, filename, is_csv=False):
    """Create Excel file with pivot sheets for detected columns."""
    try:
        output = io.BytesIO()
        wb = Workbook()
        ws_main = wb.active
        ws_main.title = "MainData"

        for r in dataframe_to_rows(df, index=False, header=True):
            ws_main.append(r)

        for col in ws_main.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = col[0].column_letter
            ws_main.column_dimensions[col_letter].width = min(max_length + 2, 100)

        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        row_fill1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        row_fill2 = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws_main[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row):
            for cell in row:
                cell.border = thin_border
                if ws_main.row_dimensions[cell.row].index % 2 == 0:
                    cell.fill = row_fill1
                else:
                    cell.fill = row_fill2

        pivot_sheets = {
            'BPartyData': ('B-Party', 'Total Calls', ['Name', 'CNIC', 'Address', 'CRO by Mobile', 'CRO by CNIC']),
            'APartyData': ('A-Party', 'Total Calls', ['Name', 'CNIC', 'Address']),
            'IMEISummary': ('IMEI', 'Count of IMEI', []),
            'LocationPivot': ('SiteLocation', 'Total Calls', []),
            'CallTypeSummary': ('Call Type', 'Count of Call Type', [])
        }

        for sheet_name, (column, count_name, optional_columns) in pivot_sheets.items():
            if column not in df.columns:
                continue
            ws = wb.create_sheet(sheet_name)
            try:
                if sheet_name in ['BPartyData', 'LocationPivot']:
                    pivot_summary = df.groupby(column).size().reset_index(name='Total Calls')
                    if sheet_name == 'BPartyData':
                        all_b_party_numbers = df[column].dropna().unique()
                        pivot_summary = pd.DataFrame(all_b_party_numbers, columns=[column])
                        pivot_summary['Total Calls'] = pivot_summary[column].map(df[column].value_counts())
                        for col in optional_columns:
                            if col in df.columns:
                                temp_df = df[[column, col]].drop_duplicates(column).dropna(subset=[column])
                                pivot_summary = pd.merge(pivot_summary, temp_df, on=column, how='left')
                    pivot_summary = pivot_summary.sort_values(by='Total Calls', ascending=False)
                else:
                    pivot_summary = df.groupby(column).size().reset_index(name=count_name)
                    for col in optional_columns:
                        if col in df.columns:
                            temp_df = df[[column, col]].drop_duplicates(column).dropna(subset=[column])
                            pivot_summary = pd.merge(pivot_summary, temp_df, on=column, how='left')
                    pivot_summary = pivot_summary.sort_values(by=count_name, ascending=False)

                if pivot_summary.empty:
                    pivot_summary = pd.DataFrame([[f"No {column} Data", 0]], columns=[column, count_name])
                for r in dataframe_to_rows(pivot_summary, index=False, header=True):
                    ws.append(r)
                for col in ws.columns:
                    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 100)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = thin_border
                        if ws.row_dimensions[cell.row].index % 2 == 0:
                            cell.fill = row_fill1
                        else:
                            cell.fill = row_fill2
            except Exception as e:
                st.error(f"Error creating {sheet_name} sheet: {e}")
                ws.append([f"Error processing {column} data"])
                ws.append([str(e)])
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
                for cell in ws[2]:
                    cell.border = thin_border
                    cell.fill = row_fill1

        default_sheet_name = wb.sheetnames[0] if wb.sheetnames else None
        if default_sheet_name and default_sheet_name != "MainData":
            wb.remove(wb[default_sheet_name])

        wb.save(output)
        output.seek(0)
        suffix = "_CSV" if is_csv else ""
        return output, f"Punjab_Cyber_Analyzer_Pivot{suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    except Exception as e:
        st.error(f"Excel generation error for {filename}: {str(e)}")
        return None, filename

def create_pdf(df, filename):
    """Create PDF file."""
    try:
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=72, bottomMargin=72)
        
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            name='CyberTitle',
            fontName='Courier-Bold',
            fontSize=16,
            textColor=colors.HexColor('#00FF00'),
            alignment=1,
            spaceBefore=15,
            spaceAfter=15,
            wordWrap='CJK'
        ))
        styles.add(ParagraphStyle(
            name='CyberHeader',
            fontName='Courier-Bold',
            fontSize=10,
            textColor=colors.black,
            backColor=colors.white,
            alignment=1,
            spaceBefore=5,
            spaceAfter=5,
            wordWrap='CJK'
        ))
        styles.add(ParagraphStyle(
            name='CyberData',
            fontName='Courier',
            fontSize=8,
            textColor=colors.black,
            leading=9,
            spaceBefore=0,
            spaceAfter=0,
            wordWrap='CJK'
        ))
        styles.add(ParagraphStyle(
            name='CyberMeta',
            fontName='Courier',
            fontSize=10,
            textColor=colors.black,
            spaceBefore=8,
            spaceAfter=8
        ))
        
        elements = [
            Paragraph("⚡ PUNJAB CYBER ANALYZER REPORT ⚡", styles['CyberTitle']),
            Spacer(1, 15),
            Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['CyberMeta']),
            Paragraph(f"Records: {len(df):,}", styles['CyberMeta']),
            Paragraph(f"Source: {filename if filename else 'Unknown'}", styles['CyberMeta']),
            Spacer(1, 25)
        ]
        
        if df is not None and not df.empty:
            total_width = 523
            num_cols = len(df.columns)
            col_width = total_width / num_cols if num_cols > 0 else total_width
            main_data = [[Paragraph(f"<b>{col}</b>", styles['CyberHeader']) for col in df.columns]]
            for _, row in df.fillna('').astype(str).head(100).iterrows():
                data_row = [Paragraph(row[col][:30], styles['CyberData']) for col in df.columns]
                main_data.append(data_row)
            main_table = Table(main_data, colWidths=[col_width] * num_cols, repeatRows=1)
            main_table.setStyle(PDFTableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.white),
                ('TEXTCOLOR', (0,0), (-1,-1), colors.black),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 3),
                ('RIGHTPADDING', (0,0), (-1,-1), 3),
                ('TOPPADDING', (0,0), (-1,-1), 3),
                ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.lightgrey])
            ]))
            elements.append(main_table)
        
        doc.build(elements)
        pdf_buffer.seek(0)
        return pdf_buffer, f"Punjab_Cyber_Analyzer_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    except Exception as e:
        st.error(f"PDF generation error: {e}")
        return None, filename

def login_page():
    st.markdown("""
    <div style="color:#00FF00; font-size:1.5rem; text-align:center; background-color:#001F00; border:2px solid #00FF00; padding:10px; margin-bottom:10px; font-weight:bold; text-shadow:0 0 10px #00FF00;">
    PUNJAB POLICE CYBER CELL LOGIN GATEWAY<br>
    <span style='font-size:1rem;'>INSPIRATION OF DIGITAL POWER AND Punjab Cyber Analyzer LEGACY</span><br>
    <span style='font-size:0.9rem;'>"We see through lies. We decode truth. We control digital warfare."</span>
    </div>
    
    <div class="whatsapp-container">
        <a href="https://wa.me/923309653269" target="_blank" class="whatsapp-link">
            ⚡ WHATSAPP ME: <span style="text-shadow:0 0 8px #00FF00;">03309653269</span> ⚡
        </a>
    </div>
    
    <div class="notice-container">
        <div class="notice-text">
            USE LAPTOP OR PC FOR BETTER VIEW
        </div>
    </div>
    
    <div class="email-container">
        <a href="mailto:punjabcyberanalyzer@gmail.com" target="_blank" class="email-link">
            📧 CONTACT US: <span style="text-shadow:0 0 8px #00FF00;">punjabcyberanalyzer@gmail.com</span> 📧
        </a>
    </div>
    """, unsafe_allow_html=True)

    username = st.text_input("OPERATOR ID").strip().upper()
    password = st.text_input("PASSCODE", type="password").strip().upper()
    license = st.text_input("LICENSE KEY", type="password").strip()

    if st.button("INITIATE SYSTEM"):
        if username not in OPERATOR_IDS or password not in PASSWORDS or license not in VALID_KEYS:
            st.error("ACCESS DENIED: INVALID CREDENTIALS OR LICENSE KEY")
        else:
            key_data = load_key_data()
            current_time = datetime.now()

            if license in key_data:
                activation_date = datetime.fromisoformat(key_data[license]["activation_date"])
                if current_time > activation_date + timedelta(days=30):
                    st.error("ACCESS DENIED: LICENSE KEY HAS EXPIRED")
                    return
            else:
                key_data[license] = {
                    "activation_date": current_time.isoformat(),
                    "last_accessed": current_time.isoformat()
                }
                save_key_data(key_data)

            key_data[license]["last_accessed"] = current_time.isoformat()
            save_key_data(key_data)

            st.session_state.license_key = license
            st.session_state.logged_in = True
            st.session_state.page = "analyzer"
            st.session_state.df = None
            st.session_state.uploaded_file = None
            st.rerun()

    st.markdown("""
    <div style="color:red; text-align:center; margin-top:30px; border-top:1px solid #00FF00; padding:10px;">
        <b>TO BUY KEY: WhatsApp → 03309653269 | 03309653269</b><br>
        <span style='color:#00FF00; font-size:0.8rem;'>"Access is power. Power demands control."</span>
    </div>
    """, unsafe_allow_html=True)

def analyzer_page():
    st.success("ANALYSIS MODULE ACTIVATED")

    key_data = load_key_data()
    license_key = st.session_state.get("license_key", "")
    if license_key in key_data:
        activation_date = datetime.fromisoformat(key_data[license_key]["activation_date"])
        expiration_date = activation_date + timedelta(days=30)
        activation_str = activation_date.strftime("%d-%b-%Y")
        expiration_str = expiration_date.strftime("%d-%b-%Y")
        days_remaining = max(0, (expiration_date - datetime.now()).days)
        st.markdown(f"""
        <div class="key-info-container">
            <div class="key-info-text">
                🔑 LICENSE KEY: {license_key}<br>
                ACTIVATED ON: {activation_str}<br>
                EXPIRES ON: {expiration_str}<br>
                DAYS REMAINING: {days_remaining}
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div class="key-info-container">
            <div class="key-info-text">
                🔑 LICENSE KEY: {license_key}<br>
                STATUS: NOT ACTIVATED
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<h4 style='color:#00FF00'>📥 DOWNLOAD OPTIONS</h4>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    button_placeholders = {'excel': col1.empty(), 'csv': col2.empty(), 'pdf': col3.empty()}

    open_spreadsheet = st.checkbox("Open Spreadsheet After Download", value=False)

    st.markdown("<h4 style='color:#00FF00'>🔍 Global Search</h4>", unsafe_allow_html=True)
    search_input = st.text_input("Search by Number / IMEI / CNIC / IMSI:", key="main_table_search")
    st.markdown("---")
    uploaded_file = st.file_uploader("CHOOSE FILE TO PROCESS", type=["xlsx", "xls", "csv"])

    if uploaded_file:
        with st.spinner("Loading file..."):
            if st.session_state.df is None or st.session_state.uploaded_file != uploaded_file.name:
                df = load_file(uploaded_file)
                st.session_state.df = df
                st.session_state.uploaded_file = uploaded_file.name
            else:
                df = st.session_state.df
        
        if df is not None and not df.empty:
            try:
                if search_input:
                    pattern = re.escape(search_input.strip())
                    df = df[df.apply(lambda row: row.astype(str).str.contains(pattern, case=False, na=False).any(), axis=1)]
                    if df.empty:
                        st.warning("No records match the search criteria.")

                df = df.loc[:, ~df.columns.duplicated()]
                df = df.reset_index(drop=True)

                styled_df = df.style.set_properties(**{
                    'font-size': '12px',
                    'border-color': '#00FF00',
                    'border-style': 'solid',
                    'border-width': '2px'
                })
                st.dataframe(styled_df, use_container_width=True)

                st.markdown("<h4 style='color:#00FF00'>📊 SUMMARY SECTION</h4>", unsafe_allow_html=True)
                pivot_sections = {
                    'A-Party': ('A-Party', 'Total Calls'),
                    'B-Party': ('B-Party', 'Total Calls'),
                    'IMEI': ('IMEI', 'Count'),
                    'Call Type': ('Call Type', 'Count'),
                    'SiteLocation': ('SiteLocation', 'Total Calls')
                }

                for column, count_name in pivot_sections.items():
                    if column not in df.columns:
                        continue
                    st.markdown(f"<h5 style='color:#00FF00'>{column} Summary</h5>", unsafe_allow_html=True)
                    pivot_summary = df[[column]].copy()
                    pivot_summary = pivot_summary.loc[:, ~pivot_summary.columns.duplicated()]
                    pivot_summary = pivot_summary.reset_index(drop=True)
                    pivot_summary = pivot_summary.groupby(column).size().reset_index(name=count_name)
                    if pivot_summary.empty:
                        pivot_summary = pd.DataFrame([[f"No {column} Data", 0]], columns=[column, count_name])
                    styled_pivot = pivot_summary.style.set_properties(**{
                        'font-size': '12px',
                        'border-color': '#00FF00',
                        'border-style': 'solid',
                        'border-width': '2px'
                    })
                    st.dataframe(styled_pivot, use_container_width=True)

                excel_data, excel_filename = create_excel(df, uploaded_file.name if uploaded_file else "cdr_data.xlsx", is_csv=False)
                with button_placeholders['excel']:
                    if excel_data:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                            tmp_file.write(excel_data.getvalue())
                            tmp_file_path = tmp_file.name
                        st.download_button("📥 DOWNLOAD EXCEL OUTPUT", excel_data, file_name=excel_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        if open_spreadsheet:
                            try:
                                if platform.system() == "Windows":
                                    os.startfile(tmp_file_path)
                                elif platform.system() == "Darwin":
                                    os.system(f"open {tmp_file_path}")
                                else:
                                    os.system(f"xdg-open {tmp_file_path}")
                            except Exception as e:
                                st.warning(f"Failed to open spreadsheet: {e}")
                            finally:
                                try:
                                    os.unlink(tmp_file_path)
                                except:
                                    pass
                    else:
                        st.warning("Excel generation failed. Please check the error message above.")

                csv_data, csv_filename = create_excel(df, uploaded_file.name if uploaded_file else "cdr_data_csv.xlsx", is_csv=True)
                with button_placeholders['csv']:
                    if csv_data:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                            tmp_file.write(csv_data.getvalue())
                            tmp_file_path = tmp_file.name
                        st.download_button("📥 DOWNLOAD CSV OUTPUT AS EXCEL", csv_data, file_name=csv_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        if open_spreadsheet:
                            try:
                                if platform.system() == "Windows":
                                    os.startfile(tmp_file_path)
                                elif platform.system() == "Darwin":
                                    os.system(f"open {tmp_file_path}")
                                else:
                                    os.system(f"xdg-open {tmp_file_path}")
                            except Exception as e:
                                st.warning(f"Failed to open spreadsheet: {e}")
                            finally:
                                try:
                                    os.unlink(tmp_file_path)
                                except:
                                    pass
                    else:
                        st.warning("CSV output generation failed. Please check the error message above.")

                pdf_data, pdf_filename = create_pdf(df, uploaded_file.name if uploaded_file else "cdr_data.pdf")
                with button_placeholders['pdf']:
                    if pdf_data:
                        st.download_button("🛡️ DOWNLOAD IN PDF", pdf_data, file_name=pdf_filename, mime="application/pdf")
                    else:
                        st.warning("PDF generation failed. Please check the error message above.")

            except Exception as e:
                st.error(f"Error processing file: {e}")
        else:
            st.warning("No data loaded or file is empty.")

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "page" not in st.session_state:
    st.session_state.page = "login"
if "df" not in st.session_state:
    st.session_state.df = None
if "uploaded_file" not in st.session_state:
    st.session_state.uploaded_file = None

if not st.session_state.logged_in:
    login_page()
else:
    analyzer_page()